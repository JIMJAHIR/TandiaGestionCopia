from functools import wraps
from flask import Blueprint, redirect, render_template, request, session, url_for
import mysql.connector
import openpyxl


from app.models import detail_collection, detail_follow_up, details_collection, details_contact, list_collection_monthly, list_users, obtener_info_clientes, obtener_info_clientes_contrato, obtener_info_clientes_filtrados, register_collection_contact
from cnx import DB_CONFIG

cobranzas_bp = Blueprint('cobranzas', __name__)


def login_required(route_function):
    @wraps(route_function)
    def wrapper(*args, **kwargs):
        if 'user_name' in session and 'user_password' in session:
            return route_function(*args, **kwargs)
        else:
            return redirect(url_for('login'))
    return wrapper


@cobranzas_bp.route('/collections', methods=['GET', 'POST'])
@login_required
def collections():
    user_id = session.get('user_id')
    user_name = session.get('user_name')
    user_password = session.get('user_password')
    user_full_name = session.get('user_full_name')
    user_area = session.get('user_area')

    if request.method == 'POST':
        ruc = request.form.get('ruc')
        lista_clientes = obtener_info_clientes_filtrados(ruc)
    else:
        lista_clientes = obtener_info_clientes()

    return render_template('cobranzas/collections.html', user_id=user_id, user_name=user_name, user_password=user_password, user_full_name=user_full_name, user_area=user_area, lista_clientes=lista_clientes)


@cobranzas_bp.route('/searchCollections', methods=['GET', 'POST'])
@login_required
def searchCollections():
    user_id = session.get('user_id')
    user_name = session.get('user_name')
    user_password = session.get('user_password')
    user_full_name = session.get('user_full_name')
    user_area = session.get('user_area')

    if request.method == 'POST':
        contract_number = request.form.get('contract_number')
        cliente = obtener_info_clientes_contrato(contract_number)
    else:
        cliente = [""]

    return render_template('cobranzas/searchCollections.html', user_id=user_id, user_name=user_name, user_password=user_password, user_full_name=user_full_name, user_area=user_area, cliente=cliente)


@cobranzas_bp.route('/monthlyCollections', methods=['GET', 'POST'])
@login_required
def monthlyCollections():
    user_id = session.get('user_id')
    user_name = session.get('user_name')
    user_password = session.get('user_password')
    user_full_name = session.get('user_full_name')
    user_area = session.get('user_area')

    list_collection = list_collection_monthly()
    users = list_users()

    return render_template('cobranzas/monthlyCollections.html', user_id=user_id, user_name=user_name, user_password=user_password, user_full_name=user_full_name, user_area=user_area, list_collection=list_collection, users=users)

@cobranzas_bp.route('/cargar_new_client', methods=['GET', 'POST'])
@login_required
def cargar_new_client():
    user_id = session.get('user_id')
    user_name = session.get('user_name')
    user_password = session.get('user_password')
    user_full_name = session.get('user_full_name')
    user_area = session.get('user_area')

    list_collection = list_collection_monthly()
    users = list_users()

    return render_template('cobranzas/cargar.html', user_id=user_id, user_name=user_name, user_password=user_password, user_full_name=user_full_name, user_area=user_area, list_collection=list_collection, users=users)


@cobranzas_bp.route('/endCollection', methods=['GET', 'POST'])
@login_required
def endTracking():

    if request.method == 'POST':
        collection_monthly_id = request.form.get('id_collection_A')

        conn = mysql.connector.connect(**DB_CONFIG)
        cursor = conn.cursor()
        cursor.execute(
            "UPDATE collection_monthly SET status = 'FINALIZADO' WHERE collection_monthly_id = %s", [collection_monthly_id])
        conn.commit()
        cursor.close()
        conn.close()

    return redirect(url_for('cobranzas.monthlyCollections'))


@cobranzas_bp.route('/detailCollection', methods=['GET', 'POST'])
@login_required
def detailCollection():
    user_id = session.get('user_id')
    user_name = session.get('user_name')
    user_password = session.get('user_password')
    user_full_name = session.get('user_full_name')
    user_area = session.get('user_area')

    id_collection = request.args.get('collection')
    details = detail_collection(id_collection)

    return render_template('cobranzas/detail_Collection.html', user_id=user_id, user_name=user_name, user_password=user_password, user_full_name=user_full_name, user_area=user_area, details=details, id_collection=id_collection)


@cobranzas_bp.route('/detail_collection_edit', methods=['GET', 'POST'])
@login_required
def detail_follow_edit():

    if request.method == 'POST':
        collection = request.form.get('id_collection')
        client_Id = request.form.get('client_Id')
        id_details = request.form.get('id_details')
        renew_amount = request.form.get('renew_amount')
        end_date = request.form.get('end_date')
        promotion = request.form.get('promotion')
        commentry = request.form.get('commentry')

        conn = mysql.connector.connect(**DB_CONFIG)
        cursor = conn.cursor()
        cursor.execute(
            "UPDATE new_client SET renew_amount = %s, end_date = %s WHERE client_id = %s", [renew_amount, end_date, client_Id])
        conn.commit()

        cursor.execute(
            "UPDATE detail_collection_monthly SET promotion = %s, commentry = %s WHERE detail_collection_monthly_id = %s", [promotion, commentry, id_details])
        conn.commit()

        conn.close()

    return redirect(url_for('cobranzas.detailCollection', collection=collection))


@cobranzas_bp.route('/detail_collection_contact', methods=['GET', 'POST'])
@login_required
def detail_collection_contact():

    if request.method == 'POST':

        collection = request.form.get('id_collection')
        detail_id = request.form.get('detail_id')
        type_contact = request.form.get('type_contact')
        comment_contact = request.form.get('comment_contact')

        register_collection_contact(detail_id, type_contact, comment_contact)



    return redirect(url_for('cobranzas.detailCollection', collection=collection))

@cobranzas_bp.route('/detail_collection_active', methods=['GET', 'POST'])
@login_required
def detail_collection_active():

    if request.method == 'POST':
        collection = request.form.get('id_collectionA')
        ruc = request.form.get('ruc_A')
        detail_idA = request.form.get('detail_idA')
        activo = request.form.get('activo')

        if activo == 'SI':
            conn = mysql.connector.connect(**DB_CONFIG)
            cursor = conn.cursor()
            cursor.execute(
                "UPDATE detail_collection_monthly SET state_collection_monthly = 'RENOVADO' WHERE detail_collection_monthly_id = %s", [detail_idA])
            conn.commit()

            cursor.execute(
                "UPDATE company SET status_comp = 'ACTIVO' WHERE ruc = %s", [ruc])
            conn.commit()
            
            cursor.close()
            conn.close()

        if activo == 'NO':
            conn = mysql.connector.connect(**DB_CONFIG)
            cursor = conn.cursor()
            cursor.execute(
                "UPDATE detail_collection_monthly SET state_collection_monthly = 'DADO DE BAJA' WHERE detail_collection_monthly_id = %s", [detail_idA])
            conn.commit()
            
            cursor.execute(
                "UPDATE company SET status_comp = 'DADO DE BAJA' WHERE ruc = %s", [ruc])
            conn.commit()
            
            cursor.close()
            conn.close()

    return redirect(url_for('cobranzas.detailCollection',collection=collection))


@cobranzas_bp.route('/detail_collectionC', methods=['GET', 'POST'])
@login_required
def detail_collectionC():
    user_id = session.get('user_id')
    user_name = session.get('user_name')
    user_password = session.get('user_password')
    user_full_name = session.get('user_full_name')
    user_area = session.get('user_area')

    id_details = request.args.get('id_details')
    detail = details_collection(id_details)
    contact = details_contact(id_details)

    return render_template('cobranzas/details_collections.html', user_id=user_id, user_name=user_name, user_password=user_password, user_full_name=user_full_name, user_area=user_area, detail=detail, contact=contact, id_details=id_details)


@cobranzas_bp.route('/upload_collection_monthly', methods=['POST'])
@login_required
def upload_collection_monthly():

    user_id = session.get('user_id')

    if 'excelFile' not in request.files:
        return redirect(request.url)
    excel_file = request.files['excelFile']

    if excel_file.filename == '':
        return redirect(request.url)

    # Leer el archivo de Excel
    try:
        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook.active
        columna_ruc = sheet['A']
    except Exception as e:
        return f"Error al leer el archivo de Excel: {str(e)}"

    # Insertar en la tabla follow_up
    conn = mysql.connector.connect(**DB_CONFIG)
    cursor = conn.cursor()
    cursor.execute(
        "INSERT INTO collection_monthly (user_id,status) VALUES (%s,'EN PROCESO')", [user_id])
    conn.commit()

    collection_monthly_id = cursor.lastrowid

    # Insertar en la tabla detail_follow_up
    for celda_ruc in columna_ruc[1:]:
        valor_ruc = celda_ruc.value
        cursor.execute(
            "INSERT INTO detail_collection_monthly (collection_monthly_id, ruc_collection_monthly, state_collection_monthly) VALUES (%s, %s,'PENDIENTE')", (collection_monthly_id, valor_ruc))
        conn.commit()

    conn.close()

    return redirect(url_for('cobranzas.monthlyCollections'))




@cobranzas_bp.route('/upload_new_client', methods=['POST'])
@login_required
def upload_new_client():

    if 'excelFile' not in request.files:
        return redirect(request.url)
    excel_file = request.files['excelFile']

    if excel_file.filename == '':
        return redirect(request.url)

    # Leer el archivo de Excel
    try:
        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook.active
    except Exception as e:
        return f"Error al leer el archivo de Excel: {str(e)}"


    conn = mysql.connector.connect(**DB_CONFIG)
    cursor = conn.cursor()

        # Iterar sobre las filas del archivo Excel
    for row in sheet.iter_rows(min_row=2, values_only=True):
        
        (
            user_id, 
            closure_type, 
            ruc, social_name, 
            legal_representative, 
            doc_number, 
            department, 
            tax_address, 
            business_line, 
            client_name, 
            client_email, 
            client_phone, 
            plan, 
            contract_type, 
            stores_number, 
            stores_conf, 
            users_number, 
            users_conf, 
            link_contract, 
            fe, 
            buy_cd, 
            guides, 
            initial_amount_plan, 
            pending_amount, 
            invoice_number, 
            renew_amount, 
            payment_date, 
            start_date, 
            end_date, 
            sale_type, 
            estado_imp

        ) = row
        
        cursor.execute(
            "SELECT user_id FROM tandia.user WHERE user_full_name = %s", [user_id])
        user = cursor.fetchone()[0]  # Obtener el resultado de la consulta

        # Insertar en la tabla new_client
        cursor.execute(
            """
            INSERT INTO new_client (
                user_id, closure_type, ruc, social_name, legal_representative, doc_number, department, 
                tax_address, business_line, client_name, client_email, client_phone, plan, contract_type, 
                stores_number, stores_conf, users_number, users_conf, link_contract, fe, buy_cd, 
                guides, initial_amount_plan, pending_amount, invoice_number, renew_amount, payment_date, 
                start_date, end_date, sale_type, estado_imp
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 
            %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """,
            (
                user, closure_type, ruc, social_name, legal_representative, doc_number, department, 
                tax_address, business_line, client_name, client_email, client_phone, plan, contract_type, 
                stores_number, stores_conf, users_number, users_conf, link_contract, fe, buy_cd, 
                guides, initial_amount_plan, pending_amount, invoice_number, renew_amount, payment_date, 
                start_date, end_date, sale_type, estado_imp
            )
        )

    conn.commit()
    conn.close()

    return redirect(url_for('cobranzas.cargar_new_client'))



@cobranzas_bp.route('/upload_new_implemetation', methods=['POST'])
@login_required
def upload_new_implemetation():

    if 'excelFile' not in request.files:
        return redirect(request.url)
    excel_file = request.files['excelFile']

    if excel_file.filename == '':
        return redirect(request.url)

    # Leer el archivo de Excel
    try:
        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook.active
    except Exception as e:
        return f"Error al leer el archivo de Excel: {str(e)}"


    conn = mysql.connector.connect(**DB_CONFIG)
    cursor = conn.cursor()

        # Iterar sobre las filas del archivo Excel
    for row in sheet.iter_rows(min_row=2, values_only=True):
        print(f"Longitud de la fila: {len(row)}")
        print(f"Fila completa: {row}")
        
        (
            ruc,
            user_id,
            contract_number,
            pass_sun_boolean,
            comercial_name,
            comercial_address,
            comercial_phone,
            comercial_email,
            taxes,
            secundary_user,
            pass_secundary_user,
            url_cd,
            pass_cd,
            expiration_cd,
            gr_pass,
            gr_id

        ) = row
        
        cursor.execute(
            "SELECT user_id FROM tandia.user WHERE user_full_name = %s", [user_id])
        user = cursor.fetchone()[0]  # Obtener el resultado de la consulta

        # Insertar en la tabla new_client
        cursor.execute(
            """
            INSERT INTO implementation (
                ruc,user_id,contract_number,pass_sun_boolean,comercial_name,comercial_address,comercial_phone,comercial_email,
                taxes,secundary_user,pass_secundary_user,url_cd,pass_cd,expiration_cd,gr_pass,gr_id
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """,
            (
                ruc,user,contract_number,pass_sun_boolean,comercial_name,comercial_address,comercial_phone,comercial_email,
                taxes,secundary_user,pass_secundary_user,url_cd,pass_cd,expiration_cd,gr_pass,gr_id
            )
        )

    conn.commit()
    conn.close()

    return redirect(url_for('cobranzas.cargar_new_client'))


@cobranzas_bp.route('/upload_new_Company', methods=['POST'])
@login_required
def upload_new_Company():

    if 'excelFile' not in request.files:
        return redirect(request.url)
    excel_file = request.files['excelFile']

    if excel_file.filename == '':
        return redirect(request.url)

    # Leer el archivo de Excel
    try:
        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook.active
    except Exception as e:
        return f"Error al leer el archivo de Excel: {str(e)}"


    conn = mysql.connector.connect(**DB_CONFIG)
    cursor = conn.cursor()

        # Iterar sobre las filas del archivo Excel
    for row in sheet.iter_rows(min_row=2, values_only=True):
        print(f"Longitud de la fila: {len(row)}")
        print(f"Fila completa: {row}")
        
        (
            user_id,	
            ruc,	
            contract_number,	
            support_email,	
            support_pass,	
            status_train,
        ) = row
        
        cursor.execute(
            "SELECT user_id FROM tandia.user WHERE user_full_name = %s", [user_id])
        user = cursor.fetchone()[0]  # Obtener el resultado de la consulta

        # Insertar en la tabla new_client
        cursor.execute(
            """
            INSERT INTO company (user_id, ruc, contract_number, support_email, support_pass, status_train
            ) VALUES (%s, %s, %s, %s, %s, %s)
            """,
            (
                user, ruc, contract_number, support_email, support_pass, status_train
            )
        )

    conn.commit()
    conn.close()

    return redirect(url_for('cobranzas.cargar_new_client'))
